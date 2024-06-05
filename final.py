import pandas as pd
import pyodbc
from openpyxl import Workbook, load_workbook
import os
from dotenv import load_dotenv
load_dotenv()


#データベース接続
SERVER = os.getenv('SERVER')
DATABASE = os.getenv('DATABASE')
DB_USER = os.getenv('DB_USER')
DB_PASSWORD = os.getenv('DB_PASSWORD')

connectionString = f'DRIVER={{SQL Server}};SERVER={SERVER};DATABASE={DATABASE};UID={DB_USER};PWD={DB_PASSWORD}'
conn = pyodbc.connect(connectionString)

cursor = conn.cursor() 
conn.commit()


#テーブル作成
cursor.execute("CREATE TABLE newmovies(id int IDENTITY(1,1), Release_Date date, Title varchar(1000), Overview varchar(2000),Popularity varchar(1000),Vote_Count int ,Vote_Average int ,Original_Language  varchar(1000) ,Genre varchar(1000) ,Poster_Url varchar(1000))")
conn.commit()


#CSV 読み込む
#作成したテーブルに CSV データを入れる
df = pd.read_csv('./mymoviedb.csv',lineterminator='\n')
for row in df.itertuples():
    cursor.execute("INSERT INTO newmovies(Release_Date, Title, Overview,Popularity,Vote_Count,Vote_Average,Original_Language,Genre,Poster_Url) values(?,?,?,?,?,?,?,?,?)",row.Release_Date,row.Title,row.Overview,row.Popularity,row.Vote_Count,row.Vote_Average,row.Original_Language,row.Genre,row.Poster_Url)


# conn.commit()

#データベースからデータ取得
df = pd.read_sql('SELECT * FROM newmovies',conn)


#言語ごとの映画の数を出力する
print(df.groupby(["Original_Language"])["Original_Language"].count())

#ジャンルごとに平均値を出す
print(df.groupby(["Genre"])["Vote_Average"].mean())

#分析結果を Excel で保存
import openpyxl
wb = Workbook()
ws_1 = wb.active
ws_1.title = "Language Counts"
ws_2 = wb.create_sheet(title="Genre Ratings")

#Original Language
data1 = df.groupby(["Original_Language"])["Original_Language"].count()

ws_1.cell(1,1,"Original Language")
ws_1.cell(1,2,"Count")

count = 1
for country, number in data1.items():
    ws_1.cell(count +1, 1,country)
    ws_1.cell(count +1, 2,number)
    count +=1


#Genre Ratings
data2 = df.groupby(["Genre"])["Vote_Average"].mean()
ws_2.cell(1,1,"Genre")
ws_2.cell(1,2,"Vote_Average")

count=1
for genre, average in data2.items():
    ws_2.cell(count +1,1,genre)
    ws_2.cell(count +1,2,average)
    count +=1


wb.save("final.xlsx")
wb.close()