from openpyxl import Workbook, load_workbook
#wb = Workbook()
#wb = load_workbook

# wb = load_workbook("excel.xlsx")
# ws = wb.active
# print(ws)

# ws= wb.create_sheet(title='newsheet')
# ws["A1"]=1
# ws.cell(1,1,1)
# ws["B1"]=2
# ws.cell(1,2,2)
# ws["A2"]=3
# ws.cell(2,1,3)
# ws["B2"]=4
# ws.cell(2,2,4)


# row_data = [1,2,3,4]
# ws.append(row_data)



# wb.save("test.xlsx")

# print(ws['A1'].value)

# for row in ws.iter_rows(min_row=1, min_col=1,max_row=2,max_col=2):
#     for cell in row:
#         print(cell.value)





wb = Workbook()
ws = wb.create_sheet(title="TestSheet")


count = 1
for i in range(2,6):
    for j in range(2,6):
        ws.cell(i,j,count)
        count += 1



sum = 0
for row in ws["B2:E5"]:
    for cell in row:
        count -= 1
        cell.value=count
        sum+=count
print(sum/count)


input_ws=wb.create_sheet("DataInput")


for i in range(1,11):
    input_ws.cell(i,1,i)

for i in range(1,11):
    cell_data=input_ws.cell(i,1).value
    ws.cell(i+1,6,cell_data+10)


wb.save("ExcelResult_saito.xlsx")



